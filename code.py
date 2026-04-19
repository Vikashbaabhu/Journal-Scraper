
import re, os, time, requests, pandas as pd
from urllib.parse import urlencode, quote
from tqdm import tqdm

# ── Config ────────────────────────────────────────────────────────────────────

MAILTO    = "rvikashbaabhu@gmail.com"   
FROM_YEAR = 2021
S2_THRESHOLD = 2000

JOURNALS = {
    "American Economic Review":         "0002-8282",
    "Quarterly Journal of Economics":   "0033-5533",
    "Science":                          "0036-8075",
    "Nature":                           "0028-0836",
    "Science Advances":                 "2375-2548",
    "Nature Human Behaviour":           "2397-3374",
    "PNAS":                             "0027-8424",
    "American Journal of Sociology":    "0002-9602",
    "American Sociological Review":     "0003-1224",
    "Organization Science":             "1047-7039",
    "Management Science":               "0025-1909",
    "Administrative Science Quarterly": "0001-8392",
    "Strategic Management Journal":     "0143-2095",
}

NON_RESEARCH_RE = re.compile(
    r"^front matter|editorial board|^acknowledgment|^acknowledgements?"
    r"|^index$|\bindex\b.*\d{4}|report of.*auditor|^corrigendum"
    r"|^erratum|^retraction|^table of contents|^issue information"
    r"|^cover image|^in this issue|^news in brief|^research highlights?"
    r"|^seven days|^world view|^obituary|^correspondence\s*$"
    r"|^letters?\s*$|^books? (in brief|received)|^diary",
    re.IGNORECASE,
)

CROSSREF_BASE   = "https://api.crossref.org/works"
OPENALEX_BASE   = "https://api.openalex.org/works"
SEMSCHOLAR_BASE = "https://api.semanticscholar.org/graph/v1/paper"

ROWS_PER_PAGE = 100
SLEEP_CR      = 0.12
SLEEP_OA      = 0.12
SLEEP_S2      = 1.1

CHECKPOINT_DIR = "checkpoints"
os.makedirs(CHECKPOINT_DIR, exist_ok=True)

session = requests.Session()
session.headers.update({"User-Agent": f"AcademicScraper/9.0 (mailto:{MAILTO})"})

# ── CrossRef: cursor pagination ───────────────────────────────────────────────

def cr_page(issn, cursor):
    """Fetch one page from CrossRef using cursor pagination."""
    r = session.get(CROSSREF_BASE, params={
        "filter":  f"issn:{issn},from-pub-date:{FROM_YEAR}",
        "select":  "DOI,title,author,abstract,published,container-title,type",
        "rows":    ROWS_PER_PAGE,
        "mailto":  MAILTO,
        "cursor":  cursor,
    }, timeout=30)
    r.raise_for_status()
    return r.json()


def cr_fetch_all(name, issn):
  
    papers = []
    cursor = "*"

    data  = cr_page(issn, cursor)
    total = data["message"]["total-results"]
    print(f"  {name}: {total:,} items in CrossRef")
    pbar  = tqdm(total=total, desc="    Fetching", unit="papers", ncols=80)

    while True:
        items = data["message"].get("items", [])

        # Exit 1: empty page
        if not items:
            break

        papers.extend(items)
        pbar.update(len(items))

        # Exit 2: fetched everything CrossRef reports
        if len(papers) >= total:
            break

        # Exit 3: partial page = last page
        if len(items) < ROWS_PER_PAGE:
            break

        # Exit 4: no next cursor
        nxt = data["message"].get("next-cursor")
        if not nxt:
            break
        cursor = nxt
        time.sleep(SLEEP_CR)
        data = cr_page(issn, cursor)

    pbar.close()
    print(f"    → Retrieved {len(papers):,} raw records")
    return papers, total

# ── Parsers ───────────────────────────────────────────────────────────────────

JATS = re.compile(r"<[^>]+>")

def clean(t):
    return re.sub(r"\s+", " ", JATS.sub(" ", t or "")).strip()

def fmt_authors(lst):
    out = []
    for a in lst or []:
        f, g = a.get("family", ""), a.get("given", "")
        out.append(f"{f}, {g[0]}." if f and g else f or "")
    return "; ".join(x for x in out if x)

def get_year(item):
    for k in ("published", "published-print", "published-online"):
        dp = (item.get(k) or {}).get("date-parts", [[]])[0]
        if dp:
            return str(dp[0])
    return ""

def is_junk(title):
    return not title or bool(NON_RESEARCH_RE.search(str(title).strip()))

def parse_item(item, jname):
    return {
        "journal":  jname,
        "title":    " ".join(item.get("title") or [""]).strip(),
        "authors":  fmt_authors(item.get("author", [])),
        "year":     get_year(item),
        "doi":      (item.get("DOI") or "").strip(),
        "abstract": clean(item.get("abstract", "")),
        "keywords": "",
    }

# ── OpenAlex batch enrichment ─────────────────────────────────────────────────

def rebuild_abstract(aii):
    if not aii:
        return ""
    mx = max((p for ps in aii.values() for p in ps), default=-1)
    if mx < 0:
        return ""
    tok = [""] * (mx + 1)
    for w, ps in aii.items():
        for p in ps:
            if 0 <= p <= mx:
                tok[p] = w
    return " ".join(t for t in tok if t)

def openalex_enrich(dois):
    
    out, SZ = {}, 50
    batches  = list(range(0, len(dois), SZ))
    iterator = tqdm(batches, desc="    OpenAlex", unit="batch", ncols=80) \
               if len(dois) > 200 else batches
    for i in iterator:
        chunk = dois[i:i + SZ]
        pipe  = "|".join(f"https://doi.org/{d}" for d in chunk)
        try:
            r = session.get(OPENALEX_BASE, params={
                "filter":   f"doi:{pipe}",
                "select":   "doi,abstract_inverted_index,keywords,concepts",
                "per-page": SZ,
                "mailto":   MAILTO,
            }, timeout=30)
            r.raise_for_status()
            for w in r.json().get("results", []):
                key = (w.get("doi") or "").replace("https://doi.org/", "").lower()
                if not key:
                    continue
                ab  = rebuild_abstract(w.get("abstract_inverted_index"))
                kws = [k.get("display_name", "") for k in (w.get("keywords") or [])]
                if not kws:
                    kws = [c.get("display_name", "") for c in (w.get("concepts") or [])
                           if c.get("score", 0) >= 0.3]
                out[key] = {
                    "abstract": ab,
                    "keywords": "; ".join(k for k in kws if k),
                }
        except Exception as e:
            print(f"\n    OpenAlex batch error (i={i}): {e}")
        time.sleep(SLEEP_OA)
    return out

# ── Semantic Scholar fallback ─────────────────────────────────────────────────

def s2_abstract(doi):
    try:
        r = session.get(f"{SEMSCHOLAR_BASE}/DOI:{doi}",
                        params={"fields": "abstract"}, timeout=15)
        if r.status_code == 200:
            return r.json().get("abstract") or ""
    except Exception:
        pass
    return ""

# ── Checkpoint helpers ────────────────────────────────────────────────────────

def ckpt_path(name):
    return os.path.join(CHECKPOINT_DIR, f"{name.replace(' ', '_')}.csv")

def load_valid_checkpoint(name, issn):
    path = ckpt_path(name)
    if not os.path.exists(path):
        return None
    try:
        probe  = cr_page(issn, "*")
        cr_tot = probe["message"]["total-results"]
        saved  = pd.read_csv(path, dtype=str).fillna("")
        if len(saved) >= cr_tot * 0.50:
            print(f"  {name}: checkpoint OK ({len(saved):,} papers) — skipping")
            return saved
        print(f"  {name}: stale checkpoint ({len(saved):,} vs {cr_tot:,}) — re-fetching")
        os.remove(path)
    except Exception:
        pass
    return None

# ── Per-journal pipeline ──────────────────────────────────────────────────────

def process_journal(name, issn):
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"{'='*60}")

    saved = load_valid_checkpoint(name, issn)
    if saved is not None:
        return saved

    # Fetch
    raw, cr_total = cr_fetch_all(name, issn)

    # Parse & filter
    rows   = [parse_item(item, name) for item in raw]
    before = len(rows)
    rows   = [r for r in rows if not is_junk(r["title"])]
    rows   = [r for r in rows if r["year"] and int(r["year"]) >= FROM_YEAR]
    print(f"  Filtered {before - len(rows):,} non-research → {len(rows):,} research papers")

    # OpenAlex enrichment
    dois = [r["doi"] for r in rows if r["doi"]]
    print(f"  OpenAlex enrichment for {len(dois):,} DOIs…")
    oa = openalex_enrich(dois)
    for row in rows:
        d = oa.get(row["doi"].lower(), {})
        if not row["abstract"] and d.get("abstract"):
            row["abstract"] = d["abstract"]
        if d.get("keywords"):
            row["keywords"] = d["keywords"]

    # Semantic Scholar fallback for small journals only
    missing = [r for r in rows if not r["abstract"] and r["doi"]]
    if missing and cr_total <= S2_THRESHOLD:
        print(f"  Semantic Scholar fallback for {len(missing):,} abstracts…")
        for row in tqdm(missing, desc="    S2", ncols=80):
            row["abstract"] = s2_abstract(row["doi"])
            time.sleep(SLEEP_S2)
    elif missing:
        print(f"  Skipping S2 for large journal — {len(missing):,} items without "
              f"abstracts are typically news/views pieces with no abstract anywhere")

    # Save
    df = pd.DataFrame(rows, columns=[
        "journal", "title", "authors", "year", "doi", "abstract", "keywords"
    ])
    ab = (df["abstract"].fillna("") != "").sum()
    kw = (df["keywords"].fillna("") != "").sum()
    print(f"  ✓ {len(df):,} papers | abstracts: {ab:,} ({ab/max(len(df),1):.0%}) "
          f"| keywords: {kw:,} ({kw/max(len(df),1):.0%})")
    df.to_csv(ckpt_path(name), index=False, encoding="utf-8-sig")
    print(f"  Checkpoint saved → {ckpt_path(name)}")
    return df

# ── Final output ──────────────────────────────────────────────────────────────

def save_outputs(df):
    df.to_csv("papers_dataset.csv", index=False, encoding="utf-8-sig")
    print("\nSaved → papers_dataset.csv")

    from openpyxl.styles import Alignment, Font, PatternFill
    with pd.ExcelWriter("papers_dataset.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Papers")
        ws = writer.sheets["Papers"]
        ws.freeze_panes = "A2"
        hfill = PatternFill("solid", fgColor="2F5496")
        hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for cell in ws[1]:
            cell.fill  = hfill
            cell.font  = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col, w, wrap in [
            ("A", 26, False), ("B", 65, False), ("C", 38, False),
            ("D",  6, False), ("E", 32, False), ("F", 80, True), ("G", 55, False),
        ]:
            ws.column_dimensions[col].width = w
            if wrap:
                for cell in ws[col][1:]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 60
    print("Saved → papers_dataset.xlsx")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    frames = []
    for name, issn in JOURNALS.items():
        try:
            frames.append(process_journal(name, issn))
        except Exception as e:
            print(f"\n  ERROR on {name}: {e}")
            import traceback; traceback.print_exc()

    df = pd.concat(frames, ignore_index=True)
    df = df[pd.to_numeric(df["year"], errors="coerce").fillna(0) >= FROM_YEAR].reset_index(drop=True)

    print(f"\n{'='*60}")
    print(f"  FINAL SUMMARY")
    print(f"{'='*60}")
    print(f"  Total papers : {len(df):,}")
    print(f"  Journals     : {df['journal'].nunique()}")
    print(f"  Year range   : {df['year'].min()} – {df['year'].max()}")
    print(f"  With abstract: {(df['abstract'].fillna('')!='').sum():,} "
          f"({(df['abstract'].fillna('')!='').mean():.1%})")
    print(f"  With keywords: {(df['keywords'].fillna('')!='').sum():,} "
          f"({(df['keywords'].fillna('')!='').mean():.1%})")
    print(f"\n  {'Journal':<40} {'Papers':>7}  {'Abs':>7}  {'Kw':>7}")
    print(f"  {'-'*65}")
    for j in df["journal"].unique():
        s  = df[df["journal"] == j]
        ab = (s["abstract"].fillna("") != "").sum()
        kw = (s["keywords"].fillna("") != "").sum()
        print(f"  {j:<40} {len(s):>7,}  "
              f"{ab:>5,} ({ab/len(s):.0%})  {kw:>5,} ({kw/len(s):.0%})")

    save_outputs(df)

if __name__ == "__main__":
    main()